"""A simple API to store/load python objects to/from spreadsheets

Limitations
-----------
Currently only classes with 51 or less attributes are supported

Examples
--------
### Store some animal instances in a spreadsheet called 'animals.xlsx'
```
from ezspreadsheet import Spreadsheet

class Animal():
    def __init__(self, name:str, conservation_status:str):
        self.name = name
        self.conservation_status = conservation_status

leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

philippine_eagle = Animal('Philippine Eagle', 'Threatened')

with Spreadsheet('animals.xlsx', Animal) as output_sheet:
    output_sheet.store(leopard_gecko, philippine_eagle)
```

### Store a list of instances into a spreadsheet called 'users.xlsx'
```
from ezspreadsheet import Spreadsheet

import random
import string
from dataclasses import dataclass

@dataclass
class User():
    Name:str
    Age:int
    Weight:int
    Family: list # Note that Iterables will be flattened to a string with newline seperators

instances = []
ranstring = lambda: ''.join(random.choices(string.ascii_uppercase, k=10)) # Generates a random 10 character string
for i in range(1000):
    instances.append(User(ranstring(), random.randint(12,100), random.randint(75,400), [ranstring(), ranstring(), ranstring()]))

with Spreadsheet('users.xlsx', User) as output_sheet:
    output_sheet.store(instances)
```
"""
import logging                               # Used to log data for debugging
import datetime                              # Used to validate type assertions for datetime instances
from collections import namedtuple
from typing import Any, Union, Iterable      # Used for type hinting and type assertions

# Third party dependencies
import colored                               # Colours terminal output for emphasis
from openpyxl import Workbook, load_workbook # Used to open and operate with xlsx files
from openpyxl.styles import Font, Alignment  # Used to pretty output to files


class Spreadsheet():
    """A class that takes in instances of objects and serializes them to xlsx files

    Parameters
    ----------
    file_name : (str)
        The name of the .xlsx file that will be saved out (extension can be included or excluded)

    class_identifier : (object or bool)
        The class object for instances you want to store, see example(s) for details
        If not specified (left as False), it's assumed you only want to load values

    Raises
    ------
    ValueError

        In two cases:

            1. If instances provided to Spreadsheet.store() do not match type used to construct Spreadsheet instance
            2. If class provided has more than 51 attributes (see limitations section of docs for details)

    Examples
    --------
    #### Store some animal instances in a spreadsheet called 'animals.xlsx'
    ```
    from ezspreadsheet import Spreadsheet

    class Animal():
        def __init__(self, name:str, conservation_status:str):
            self.name = name
            self.conservation_status = conservation_status

    leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

    philippine_eagle = Animal('Philippine Eagle', 'Threatened')

    with Spreadsheet('animals.xlsx', Animal) as output_sheet:
        output_sheet.store(leopard_gecko, philippine_eagle)
    ```
    """
    def __init__(self, file_name:str, class_identifier:object=False):
        self.file_name = file_name
        self.workbook = None
        self.worksheet = None
        self.class_identifier = class_identifier


        # Make sure filename has .xlsx extension TODO: Remove
        if not file_name.endswith(".xlsx"):
            logging.debug(f"added .xlsx to {file_name}")
            file_name += ".xlsx"

        if class_identifier:
            # Get all attributes of class defined in __init__
            self.class_attributes = class_identifier.__init__.__code__.co_varnames[1::] # Skip the self
            if len(self.class_attributes) > 51:
                raise ValueError(f"Provided class {class_identifier.__name__} has more than 51 attributes")


    def __enter__(self):
        """Entrypoint for the context manager

        Returns
        -------
        Spreadsheet
            Reference to self
        """
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.page_setup.fitToWidth = 1
        return self


    def __exit__(self, exc_type, exc_value, traceback):
        """Exitpoint for the context manager

        Returns
        -------
        bool
            True if the context manager ran into no issues saving files
        """
        if exc_type is None and exc_value is None:
            try:
                self.workbook.save(self.file_name)
                print(f"{self.file_name} successfully saved")
            except PermissionError:
                input(f"{colored.fg(1)}File {self.file_name} is currently open{colored.fg(15)}\nPlease close it and hit enter to save file: ")
                self.workbook.save(self.file_name)
            return True

        else:
            print(f"{colored.fg(1)}Ran into exception {exc_type}, with value {exc_value} here is traceback{colored.fg(15)}")
            return False


    def _add_row_to_spreadsheet(self, data:list, row:int, style:Font = False, readable:bool = False):
        """Take in some data, and an int for a row and add that data to that row

        Parameters
        ----------
        data : list
            The data you want to store in the row

        row : int
            The index of the row to store the data to

        style : Font, optional
            If you want to supply custom formatting for the row, by default False

        readable : bool
            If True iterable attributes are written as readable values instead of directly storing iterables, by default False

        Note
        ----

        - iterables stored while readable == true cannot be deserialized to their original type
        """
        # The value that will be converted using chr() for column identifiers i.e. A1 B1 etc.
        column_identifier = 65  # Initialize to ord() value of 'A'

        for value in data:
            if column_identifier == 91:  # Roll over to Ax column identifiers from x column identifiers
                label = f"AA{row}"
            elif column_identifier > 91:  # If beyond Z in column identifiers
                label = f"A{chr(column_identifier-26)}{row}"
            else:  # If before or at Z in column identifiers
                label = f"{chr(column_identifier)}{row}"
            logging.debug(f"{value} will be written to {label}")

            # Apply styles if specified
            if style:
                self.worksheet[label].font = style

            # Add value to worksheet
            if type(value) not in [str, int, float, datetime.datetime]:
                if type(value) == dict and readable:
                    print("Serializing dictionary in readable format") # TODO: remove
                    logging.debug("Serializing dictionary in readable format")
                    flattened_value = ""
                    for key in value:
                        flattened_value += f"- {key}: {value[key]}\n"
                    self.worksheet[label] = flattened_value
                
                elif readable:
                    # If value is an Iterable that's not a str, int or float then flatten it to a str
                    logging.debug(f"Serializing {type(value)} in readable format")
                    flattened_value = ""
                    for sub_value in value: 
                        flattened_value += f"- {str(sub_value)}\n"
                    self.worksheet[label] = flattened_value

                else:
                    # Value is not a str, int, float or datetime object (all can be natively serialized)
                    self.worksheet[label] = str(value)
            else: # If value is a string, int, float or datetime object
                self.worksheet[label] = value

            # Apply wrap text formatting to all rows that aren't the heading
            if not row == 1: 
                self.worksheet[label].alignment = Alignment(wrapText=True)

            # Increment the column identifiers variable to move to next column letter
            column_identifier += 1


    def _get_values_from_instance(self, instance:object) -> list:
        """Get's the instance's attribute values

        Parameters
        ----------
        instance : object
            The instance to pull the attribute values from

        Returns
        -------
        list
            The values for the attributes from the instance
        """
        logging.debug(f"Attributes are {self.class_attributes}")
        values = [] # All the values of the attributes in order
        for attribute in self.class_attributes:
            logging.debug(f"Looking for attribute {attribute} found value {instance.__dict__[attribute]}")
            values.append(instance.__dict__[attribute]) 
        return values


    def store(self, *instances:Union[object, Iterable[object]], readable:bool = False):
        """Takes in instance(s) of the specified class to store

        Parameters
        ----------
        instances : (Iterable[object] or arbitrary number of isntances)
            The instances with the data you want to store

        readable : bool
            If True iterable attributes are written as readable values instead of directly storing iterables, by default False

        Notes
        -----

        - iterables stored while readable == true cannot be deserialized to their original type

        Raises
        ------
        ValueError
            If an instance is not the correct type

        Notes
        -----

        - Any methods are not serialized, only attribtues

        Examples
        --------
        #### Store some animal instances in a spreadsheet called 'animals.xlsx'
        ```
        from ezspreadsheet import Spreadsheet

        class Animal():
            def __init__(self, name:str, conservation_status:str):
                self.name = name
                self.conservation_status = conservation_status

        leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

        philippine_eagle = Animal('Philippine Eagle', 'Threatened')

        with Spreadsheet('animals.xlsx', Animal) as output_sheet:
            output_sheet.store(leopard_gecko, philippine_eagle)
        ```
        """
        print(f"Beginning to store {self.class_identifier.__name__} instances to {self.file_name}")
        current_row = 1  # The current row that the iteration is at

        # Add heading with the list of class attributes to A1
        self._add_row_to_spreadsheet(self.class_attributes, current_row, Font(bold=True, size=14))
        current_row += 1  # Increment row to start with row right after heading
        logging.debug(f"Instances are {instances}")

        # Check if instance provided is a class of correct type, or an Iterable
        for current_instance in instances:
            logging.debug(f"Instance is {str(current_instance)}")
            if isinstance(current_instance, Iterable):  # If argument is an Iterable (i.e. list, tuple etc.)
                for sub_instance in current_instance:
                    if not isinstance(sub_instance, self.class_identifier):  # Validate sub-instance is correct type
                        raise ValueError(f"Provided instance: {sub_instance} is not of type {self.class_identifier}")
                    else:
                        self._add_row_to_spreadsheet(self._get_values_from_instance(sub_instance), current_row, readable=readable)
                        current_row += 1
            elif not isinstance(current_instance, self.class_identifier):  # If argument is not correct type
                raise ValueError(f"Provided instance: {current_instance} is not of type {self.class_identifier}")
            
            else:  # If argument is a single class instance of the correct type
                logging.debug(f"Adding values from {str(current_instance)}: {self._get_values_from_instance(current_instance)}")
                self._add_row_to_spreadsheet(self._get_values_from_instance(current_instance), current_row, readable=readable)
                current_row += 1

    def _load_values(self) -> list:
        """Yields each row of values to be consumed inside self.load()

        Yields
        -------
        list
            The values for a given row
        """
        for values in self.worksheet.values:
            values = list(values)
            for index, value in enumerate(values):
                # Deserialize iterables like lists, tuples and dicts
                if type(value) == str:
                    if value.startswith("["):
                        value = value[1:-2].replace("\'", "").replace('\"', "").split(',')
                        values[index] = [v.strip() for v in value]

                    elif value.startswith("("):
                        value = value[1:-2].replace("\'", "").replace('\"', "").split(',')
                        values[index] = tuple(v.strip() for v in value)

                    elif value.startswith("{"):
                        key_value_pairs = value[1:-2].replace("\'", "").replace('\"', "").split(',')
                        result = {}
                        for pair in key_value_pairs:
                            key, value = pair.split(":")
                            key = key.strip()
                            if type(value) == str:
                                result[key] = value.strip()
                            else:
                                result[key] = value
                        values[index] = result
            yield values

    def load(self, name:str) -> tuple:
        """Loads the class, and instances stored inside Spreadsheet at self.file_name

        Parameters
        ----------
        name : str
            The name you want to assign the class that is returned

        Notes
        -----

        - if self.class_identifier is specified on Spreadsheet instantiation then that class is used instead of instantiating a new one

        Returns
        -------
        tuple
            First return value is the constructor used to create instances (class if class_identifier is specified, else namedtuple), and second all the found instances

        Notes
        -----

        - If you didn't specify a class identifier when opening the spreadsheet the returned values are namedtuples and not full class instances

        Examples
        --------
        #### Loading some stored values of the Animal class from animals.xlsx
        ```
        with Spreadsheet('animals.xlsx') as loaded_sheet:
            Animal, instances = loaded_sheet.load('Animal')

        # NOTE: Animal at this point is a namedtuple constructor, not a full python class

        print(Animal) # Prints: <class '__main__.Animal'>
        print(instances) # Prints: [Animal(name='Leopard Gecko', conservation_status='Least Concern'), Animal(name='Philippine Eagle', conservation_status='Threatened')]
        ```

        #### Loading some stored values of the Animal class from animals.xlsx with the class identifier specified
        ```
        class Animal():
            def __init__(self, name:str, conservation_status:str):
                self.name = name
                self.conservation_status = conservation_status
        
        with Spreadsheet('animals.xlsx', Animal) as loaded_sheet:
            Animal, instances = loaded_sheet.load('Animal')
    
        print(Animal) # Prints: <class '__main__.Animal'>

        for instance in instances:
            print(vars(instance)) # Since these are real class instances we can use vars()
        '''prints:
        {'name': 'Leopard Gecko', 'conservation_status': 'Least Concern'}
        {'name': 'Philippine Eagle', 'conservation_status': 'Threatened'}
        '''
        ```
        """
        self.workbook = load_workbook(self.file_name)
        self.worksheet = self.workbook.active

        values = self._load_values()

        instances = []

        if self.class_identifier: # If class was specified
            logging.debug(f"Class identifier {self.class_identifier} specified")
            next(values) # skip the attributes
            constructor = self.class_identifier
            for instance_values in values:
                instances.append(self.class_identifier(*instance_values))
        else:
            logging.debug("No class identifier specified, generating namedtuple")
            # Get attributes from first row
            constructor = namedtuple(name, next(values))

            for instance_values in values:
                instances.append(constructor._make(instance_values))

        logging.debug(f"Returning: {constructor}\n\n{instances}")
        return constructor, instances
