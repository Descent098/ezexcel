# Standard library dependencies
import os                           # Used to validate file existence, and remove test files
import random                       # Used to generate random values for test classes
import string                       # Used to filter random values for test classes
from unittest import mock           # Used to fake results for tests with inputs
from dataclasses import dataclass   # Used to save time instantiating test classes

# Internal Dependencies
from ezexcel import Spreadsheet     # Functionality being tested

# Third Party Dependencies
from openpyxl import load_workbook  # Used to load a workbook and validate values


# Create test classes

@dataclass
class User():
    # Test dataclass
    Name:str
    Age:int
    Weight:int
    Family: list # Note that Iterables will be flattened to a string with newline seperators


class Animal():
    # Test standard class
    def __init__(self, name:str, conservation_status:str):
        self.name = name
        self.conservation_status = conservation_status


def test_single_instance():
    """Validates that a single instance can be stored to a sheet"""

    # Setup test instance
    leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

    # Write test instance to spreadsheet
    with Spreadsheet('animals.xlsx', Animal) as output_sheet:
        output_sheet.store(leopard_gecko)
    
    # Validate data saved properly
    workbook = load_workbook(filename = 'animals.xlsx', data_only=True)
    sheet = workbook.active
    # Confirm heading is correct
    assert sheet["A1"].value == "name"
    assert sheet["B1"].value == "conservation_status"

    # Check the A2 and B2 values are equal to leopard_gecko
    assert sheet["A2"].value == leopard_gecko.name
    assert sheet["B2"].value == leopard_gecko.conservation_status

    os.remove("animals.xlsx")  # remove animals.xlsx



def test_multi_instance():
    """Validates that multiple single instance can be stored to a sheet"""

    # Setup test instances
    leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

    philippine_eagle = Animal('Philippine Eagle', 'Threatened')

    # Write test instances to a sheet
    with Spreadsheet('animals.xlsx', Animal) as output_sheet:
        output_sheet.store(leopard_gecko, philippine_eagle)

    # Validate data saved properly
    workbook = load_workbook(filename = 'animals.xlsx', data_only=True)
    sheet = workbook.active
    # Confirm heading is correct
    assert sheet["A1"].value == "name"
    assert sheet["B1"].value == "conservation_status"

    # Check the A2 and B2 values are equal to leopard_gecko
    assert sheet["A2"].value == leopard_gecko.name
    assert sheet["B2"].value == leopard_gecko.conservation_status
    
    # Check the A3 and B3 values are equal to philippine_eagle
    assert sheet["A3"].value == philippine_eagle.name
    assert sheet["B3"].value == philippine_eagle.conservation_status

    os.remove("animals.xlsx")  # remove animals.xlsx

def test_iterable_instances():
    """Validates that an iterable of instances can be stored to a sheet"""
    instances = []
    ranstring = lambda: ''.join(random.choices(string.ascii_uppercase, k=10)) # Generates a random 10 character string
    instances.append(User("John Doe", 20, 75, ["Abby", "Mike", "Janice"]))
    for i in range(1000):
        instances.append(User(ranstring(), random.randint(12,100), random.randint(75,400), [ranstring(), ranstring(), ranstring()]))

    with Spreadsheet('users.xlsx', User) as output_sheet:
        output_sheet.store(instances)

    # Validate data saved properly
    workbook = load_workbook(filename = 'users.xlsx', data_only=True)
    sheet = workbook.active
    # Confirm heading is correct
    assert sheet["A1"].value == "Name"
    assert sheet["B1"].value == "Age"
    assert sheet["C1"].value == "Weight"
    assert sheet["D1"].value == "Family"

    # Check the second row is equal to the first instance
    assert sheet["A2"].value == "John Doe"
    assert sheet["B2"].value == 20
    assert sheet["C2"].value == 75
    assert sheet["D2"].value == "Abby\nMike\nJanice\n"

    os.remove("users.xlsx")  # remove users.xlsx


def test_flattened_list_attributes():
    """Validates that list elements are flattened on save"""
    test_user = User("John Doe", 20, 75, ["Abby", "Mike", "Janice"])

    with Spreadsheet('users.xlsx', User) as output_sheet:
        output_sheet.store(test_user)

    # Validate data saved properly
    workbook = load_workbook(filename = 'users.xlsx', data_only=True)
    sheet = workbook.active

    # Check the list has been flattened
    assert sheet["D2"].value == "Abby\nMike\nJanice\n"

    os.remove("users.xlsx")  # remove users.xlsx


def test_unmatched_instances():
    """Validates that an error is thrown when mismatched instance types are tried to be stored"""
    # Setup test instances
    leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

    philippine_eagle = Animal('Philippine Eagle', 'Threatened')

    # Try to use mismatching instances of class (User and Animal)
    try:
        with Spreadsheet('users.xlsx', User) as output_sheet:
            output_sheet.store(leopard_gecko, philippine_eagle)
        os.remove("users.xlsx")  # remove users.xlsx
        assert False # Test fails because it allows to store a different class than provided
    except ValueError:
        assert True
