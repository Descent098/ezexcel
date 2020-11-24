# Standard library dependencies
from collections import namedtuple
import os                           # Used to validate file existence, and remove test files
import random                       # Used to generate random values for test classes
import string                       # Used to filter random values for test classes
from dataclasses import dataclass   # Used to save time instantiating test classes

# Internal Dependencies
from ezspreadsheet import Spreadsheet, _CSV_Spreadsheet, _XLSX_Spreadsheet  # Functionality being tested

# Third Party Dependencies
from openpyxl import load_workbook  # Used to load a workbook and validate values


# Create test classes

@dataclass
class User():
    # Test dataclass
    Name:str
    Age:int
    Weight:int
    Family: list # Note that Iterables will be flattened to a string with newline seperators


class Animal():
    # Test standard class
    def __init__(self, name:str, conservation_status:str):
        self.name = name
        self.conservation_status = conservation_status


# Xlxs Tests

def test_xlsx_store_single_instance():
    """Validates that a single instance can be stored to a sheet"""

    # Setup test instance
    leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

    # Write test instance to spreadsheet
    with Spreadsheet('animals.xlsx', Animal) as output_sheet:
        output_sheet.store(leopard_gecko)
    
    # Validate data saved properly
    workbook = load_workbook(filename = 'animals.xlsx', data_only=True)
    sheet = workbook.active
    # Confirm heading is correct
    assert sheet["A1"].value == "name"
    assert sheet["B1"].value == "conservation_status"

    # Check the A2 and B2 values are equal to leopard_gecko
    assert sheet["A2"].value == leopard_gecko.name
    assert sheet["B2"].value == leopard_gecko.conservation_status

    os.remove("animals.xlsx")  # remove animals.xlsx



def test_xlsx_store_multi_instance():
    """Validates that multiple single instance can be stored to a sheet"""

    # Setup test instances
    leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

    philippine_eagle = Animal('Philippine Eagle', 'Threatened')

    # Write test instances to a sheet
    with Spreadsheet('animals.xlsx', Animal) as output_sheet:
        output_sheet.store(leopard_gecko, philippine_eagle)

    # Validate data saved properly
    workbook = load_workbook(filename = 'animals.xlsx', data_only=True)
    sheet = workbook.active
    # Confirm heading is correct
    assert sheet["A1"].value == "name"
    assert sheet["B1"].value == "conservation_status"

    # Check the A2 and B2 values are equal to leopard_gecko
    assert sheet["A2"].value == leopard_gecko.name
    assert sheet["B2"].value == leopard_gecko.conservation_status
    
    # Check the A3 and B3 values are equal to philippine_eagle
    assert sheet["A3"].value == philippine_eagle.name
    assert sheet["B3"].value == philippine_eagle.conservation_status

    os.remove("animals.xlsx")  # remove animals.xlsx

def test_xlsx_store_iterable_instances():
    """Validates that an iterable of instances can be stored to a sheet"""
    instances = []
    ranstring = lambda: ''.join(random.choices(string.ascii_uppercase, k=10)) # Generates a random 10 character string
    instances.append(User("John Doe", 20, 75, ["Abby", "Mike", "Janice"]))
    for i in range(1000):
        instances.append(User(ranstring(), random.randint(12,100), random.randint(75,400), [ranstring(), ranstring(), ranstring()]))

    with Spreadsheet('users.xlsx', User) as output_sheet:
        output_sheet.store(instances)

    # Validate data saved properly
    workbook = load_workbook(filename = 'users.xlsx', data_only=True)
    sheet = workbook.active
    # Confirm heading is correct
    assert sheet["A1"].value == "Name"
    assert sheet["B1"].value == "Age"
    assert sheet["C1"].value == "Weight"
    assert sheet["D1"].value == "Family"

    # Check the second row is equal to the first instance
    assert sheet["A2"].value == "John Doe"
    assert sheet["B2"].value == 20
    assert sheet["C2"].value == 75
    assert sheet["D2"].value == "['Abby', 'Mike', 'Janice']"

    os.remove("users.xlsx")  # remove users.xlsx


def test_xlsx_flattened_list_attributes():
    """Validates that list elements are flattened on save when readable flag is set on store()"""
    test_user = User("John Doe", 20, 75, ["Abby", "Mike", "Janice"])

    with Spreadsheet('users.xlsx', User) as output_sheet:
        output_sheet.store(test_user, readable = True)

    # Validate data saved properly
    workbook = load_workbook(filename = 'users.xlsx', data_only=True)
    sheet = workbook.active

    # Check the list has been flattened
    assert sheet["D2"].value == "- Abby\n- Mike\n- Janice\n"

    os.remove("users.xlsx")  # remove users.xlsx

def test_xlsx_flattened_tuple_attributes():
    """Validates that tuple elements are flattened on save when readable flag is set on store()"""
    test_user = User("John Doe", 20, 75, ("Abby", "Mike", "Janice"))

    with Spreadsheet('users.xlsx', User) as output_sheet:
        output_sheet.store(test_user, readable = True)

    # Validate data saved properly
    workbook = load_workbook(filename = 'users.xlsx', data_only=True)
    sheet = workbook.active

    # Check the tuple has been flattened
    assert sheet["D2"].value == "- Abby\n- Mike\n- Janice\n"

    os.remove("users.xlsx")  # remove users.xlsx

def test_xlsx_flattened_dict_attributes():
    """Validates that dict elements are flattened on save when readable flag is set on store()"""
    test_user = User("John Doe", 20, 75, {"age": 21, "name": "Francis"})

    with Spreadsheet('users.xlsx', User) as output_sheet:
        output_sheet.store(test_user, readable = True)

    # Validate data saved properly
    workbook = load_workbook(filename = 'users.xlsx', data_only=True)
    sheet = workbook.active

    # Check the dict has been flattened
    assert sheet["D2"].value == "- age: 21\n- name: Francis\n"

    os.remove("users.xlsx")  # remove users.xlsx

def test_xlsx_unflattened_list_attributes():
    """Validates that list elements are flattened on save when readable flag is set to False on store()"""
    test_user = User("John Doe", 20, 75, ["Abby", "Mike", "Janice"])

    with Spreadsheet('users.xlsx', User) as output_sheet:
        output_sheet.store(test_user)

    # Validate data saved properly
    workbook = load_workbook(filename = 'users.xlsx', data_only=True)
    sheet = workbook.active

    # Check the list has not been flattened
    with Spreadsheet("users.xlsx") as loaded_sheet:
        users, instances = loaded_sheet.load("users")
    assert instances[0].Family == ["Abby", "Mike", "Janice"]

    os.remove("users.xlsx")  # remove users.xlsx

def test_xlsx_unflattened_tuple_attributes():
    """Validates that tuple elements are flattened on save when readable flag is set to False on store()"""
    test_user = User("John Doe", 20, 75, ("Abby", "Mike", "Janice"))

    with Spreadsheet('users.xlsx', User) as output_sheet:
        output_sheet.store(test_user)

    # Validate data saved properly
    workbook = load_workbook(filename = 'users.xlsx', data_only=True)
    sheet = workbook.active

    # Check the tuple has not been flattened
    with Spreadsheet("users.xlsx") as loaded_sheet:
        users, instances = loaded_sheet.load("users")
    assert instances[0].Family == ("Abby", "Mike", "Janice")

    os.remove("users.xlsx")  # remove users.xlsx


def test_xlsx_unflattened_dict_attributes():
    """Validates that dict elements are flattened on save when readable flag is set to False on store()"""
    test_user = User("John Doe", 20, 75, {"age": 21, "name": "Francis"})

    with Spreadsheet('users.xlsx', User) as output_sheet:
        output_sheet.store(test_user)

    # Validate data saved properly
    workbook = load_workbook(filename = 'users.xlsx', data_only=True)
    sheet = workbook.active

    # Check the dict has been not flattened
    with Spreadsheet("users.xlsx") as loaded_sheet:
        users, instances = loaded_sheet.load("users")
    assert instances[0].Family == {"age": "21", "name": "Francis"}

    os.remove("users.xlsx")  # remove users.xlsx


def test_xlsx_unmatched_instances():
    """Validates that an error is thrown when mismatched instance types are tried to be stored"""
    # Setup test instances
    leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

    philippine_eagle = Animal('Philippine Eagle', 'Threatened')

    # Try to use mismatching instances of class (User and Animal)
    try:
        with Spreadsheet('users.xlsx', User) as output_sheet:
            output_sheet.store(leopard_gecko, philippine_eagle)
        os.remove("users.xlsx")  # remove users.xlsx
        assert False # Test fails because it allows to store a different class than provided
    except ValueError:
        assert True


def test_xlsx_load_single_instance_no_identifier():
    """Validates that a single instance can be loaded from a sheet without a class identifier"""

    # Setup test instance
    leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

    # Write test instance to spreadsheet
    with Spreadsheet('animals.xlsx', Animal) as output_sheet:
        output_sheet.store(leopard_gecko)

    with Spreadsheet("animals.xlsx") as loaded_sheet:
        animals, instances = loaded_sheet.load("animals")

    assert animals.__name__ == "animals"
    assert type(animals) == type(namedtuple("animals", ["name", "conservation_status"]))
    assert animals._make
    assert instances[0].name == "Leopard Gecko"
    assert instances[0].conservation_status == "Least Concern"

    os.remove("animals.xlsx")  # remove animals.xlsx


def test_xlsx_load_single_instance_with_identifier():
    """Validates that a single instance can be loaded from a sheet with a class identifier"""

    # Setup test instance
    leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

    # Write test instance to spreadsheet
    with Spreadsheet('animals.xlsx', Animal) as output_sheet:
        output_sheet.store(leopard_gecko)

    with Spreadsheet("animals.xlsx", Animal) as loaded_sheet:
        animals, instances = loaded_sheet.load("animals")

    assert type(animals) == type(Animal)
    assert instances[0].name == "Leopard Gecko"
    assert instances[0].conservation_status == "Least Concern"

    os.remove("animals.xlsx")  # remove animals.xlsx


def test_xlsx_load_multi_instance_no_identifier():
    """Validates that multiple instance can be loaded from a sheet without a class identifier"""

    # Setup test instance
    leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

    philippine_eagle = Animal('Philippine Eagle', 'Threatened')

    # Write test instance to spreadsheet
    with Spreadsheet('animals.xlsx', Animal) as output_sheet:
        output_sheet.store(leopard_gecko, philippine_eagle)

    with Spreadsheet("animals.xlsx") as loaded_sheet:
        animals, instances = loaded_sheet.load("animals")

    assert animals.__name__ == "animals"
    assert type(animals) == type(namedtuple("animals", ["name", "conservation_status"]))
    assert animals._make
    assert len(instances) == 2
    assert instances[0].name == "Leopard Gecko"
    assert instances[0].conservation_status == "Least Concern"
    assert instances[1].name == "Philippine Eagle"
    assert instances[1].conservation_status == "Threatened"

    os.remove("animals.xlsx")  # remove animals.xlsx


def test_xlsx_load_multi_instance_with_identifier():
    """Validates that multiple instance can be loaded from a sheet with a class identifier"""

    # Setup test instance
    leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

    philippine_eagle = Animal('Philippine Eagle', 'Threatened')

    # Write test instance to spreadsheet
    with Spreadsheet('animals.xlsx', Animal) as output_sheet:
        output_sheet.store(leopard_gecko, philippine_eagle)

    with Spreadsheet("animals.xlsx") as loaded_sheet:
        animals, instances = loaded_sheet.load("animals")

    assert type(animals) == type(Animal)
    assert instances[0].name == "Leopard Gecko"
    assert instances[0].conservation_status == "Least Concern"
    assert instances[1].name == "Philippine Eagle"
    assert instances[1].conservation_status == "Threatened"

    os.remove("animals.xlsx")  # remove animals.xlsx


def test_xlsx_store_iterable_instances_no_identifier():
    """Validates that an iterable of instances can be loaded from a sheet with no identifier"""
    instances = []
    ranstring = lambda: ''.join(random.choices(string.ascii_uppercase, k=10)) # Generates a random 10 character string
    instances.append(User("John Doe", 20, 75, ["Abby", "Mike", "Janice"]))
    for i in range(1000):
        instances.append(User(ranstring(), random.randint(12,100), random.randint(75,400), [ranstring(), ranstring(), ranstring()]))

    with Spreadsheet('users.xlsx', User) as output_sheet:
        output_sheet.store(instances)

    with Spreadsheet("users.xlsx") as loaded_sheet:
        users, instances = loaded_sheet.load("users")

    assert users.__name__ == "users"
    assert type(users) == type(namedtuple("users", ["name", "age", "weight", "family"]))
    assert users._make
    assert instances[0].Name == "John Doe"
    assert instances[0].Age == 20
    assert instances[0].Weight == 75
    assert instances[0].Family == ["Abby", "Mike", "Janice"]
    assert len(instances) == 1001

    os.remove("users.xlsx")  # remove animals.xlsx

# csv Tests

def test_csv_store_single_instance():
    """Validates that a single instance can be stored to a sheet"""

    # Setup test instance
    leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

    # Write test instance to spreadsheet
    with Spreadsheet('animals.csv', Animal) as output_sheet:
        output_sheet.store(leopard_gecko)
    
    # Validate data saved properly

    with Spreadsheet('animals.csv') as input_sheet:
        assert type(input_sheet) == _CSV_Spreadsheet
        _, instances = input_sheet.load("Animal")
        header = input_sheet.class_attributes

    # Confirm heading is correct
    assert header[0] == "name"
    assert header[1] == "conservation_status"

    # Check the A2 and B2 values are equal to leopard_gecko
    assert instances[0].name == leopard_gecko.name
    assert instances[0].conservation_status == leopard_gecko.conservation_status

    os.remove("animals.csv")  # remove animals.csv



def test_csv_store_multi_instance():
    """Validates that multiple single instance can be stored to a sheet"""

    # Setup test instances
    leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

    philippine_eagle = Animal('Philippine Eagle', 'Threatened')

    # Write test instance to spreadsheet
    with Spreadsheet('animals.csv', Animal) as output_sheet:
        output_sheet.store(leopard_gecko, philippine_eagle)

    # Validate data saved properly
    with Spreadsheet('animals.csv') as input_sheet:
        assert type(input_sheet) == _CSV_Spreadsheet
        _, instances = input_sheet.load("Animal")
        header = input_sheet.class_attributes

    # Confirm heading is correct
    assert header[0] == "name"
    assert header[1] == "conservation_status"

    # Check the first instance values are equal to leopard_gecko
    assert instances[0].name == leopard_gecko.name
    assert instances[0].conservation_status == leopard_gecko.conservation_status
    
    # Check the A3 and B3 values are equal to philippine_eagle
    assert instances[1].name == philippine_eagle.name
    assert instances[1].conservation_status == philippine_eagle.conservation_status

    os.remove("animals.csv")  # remove animals.csv

def test_csv_store_iterable_instances():
    """Validates that an iterable of instances can be stored to a sheet"""
    test_instances = []
    ranstring = lambda: ''.join(random.choices(string.ascii_uppercase, k=10)) # Generates a random 10 character string
    test_instances.append(User("John Doe", 20, 75, ["Abby", "Mike", "Janice"]))
    for i in range(1000):
        test_instances.append(User(ranstring(), random.randint(12,100), random.randint(75,400), [ranstring(), ranstring(), ranstring()]))

    # Write test instance to spreadsheet
    with Spreadsheet('users.csv', User) as output_sheet:
        output_sheet.store(test_instances)

    # Validate data saved properly
    with Spreadsheet('users.csv') as input_sheet:
        assert type(input_sheet) == _CSV_Spreadsheet
        _, instances = input_sheet.load("users")
        header = input_sheet.class_attributes

    # Confirm heading is correct
    assert header[0] == "Name"
    assert header[1] == "Age"
    assert header[2] == "Weight"
    assert header[3] == "Family"

    # Check the second row is equal to the first instance
    assert instances[0].Name == "John Doe"
    assert instances[0].Age == 20
    assert instances[0].Weight == 75
    assert instances[0].Family == ['Abby', 'Mike', 'Janice']

    os.remove("users.csv")  # remove users.csv


def test_csv_flattened_list_attributes():
    """Validates that list elements are flattened on save when readable flag is set on store()"""
    test_user = User("John Doe", 20, 75, ["Abby", "Mike", "Janice"])

    # Write test instance to spreadsheet
    with Spreadsheet('users.csv', User) as output_sheet:
        output_sheet.store(test_user, readable = True)

    # Validate data saved properly
    with Spreadsheet('users.csv') as input_sheet:
        assert type(input_sheet) == _CSV_Spreadsheet
        _, instances = input_sheet.load("users")
        header = input_sheet.class_attributes

    # Check the list has been flattened
    assert instances[0].Family == "- Abby \t- Mike \t- Janice \t"

    os.remove("users.csv")  # remove users.csv


def test_csv_flattened_tuple_attributes():
    """Validates that tuple elements are flattened on save when readable flag is set on store()"""
    test_user = User("John Doe", 20, 75, ("Abby", "Mike", "Janice"))

    # Write test instance to spreadsheet
    with Spreadsheet('users.csv', User) as output_sheet:
        output_sheet.store(test_user, readable = True)

    # Validate data saved properly
    with Spreadsheet('users.csv') as input_sheet:
        assert type(input_sheet) == _CSV_Spreadsheet
        _, instances = input_sheet.load("users")
        header = input_sheet.class_attributes

    # Check the list has been flattened
    assert instances[0].Family == "- Abby \t- Mike \t- Janice \t"

    os.remove("users.csv")  # remove users.csv


def test_csv_flattened_dict_attributes():
    """Validates that dict elements are flattened on save when readable flag is set on store()"""
    test_user = User("John Doe", 20, 75, {"age": 21, "name": "Francis"})

    # Write test instance to spreadsheet
    with Spreadsheet('users.csv', User) as output_sheet:
        output_sheet.store(test_user, readable = True)


    # Validate data saved properly

    with Spreadsheet('users.csv') as input_sheet:
        assert type(input_sheet) == _CSV_Spreadsheet
        _, instances = input_sheet.load("users")
        header = input_sheet.class_attributes

    # Check the dict has been flattened
    assert instances[0].Family == "- age: 21 \t- name: Francis \t"

    os.remove("users.csv")  # remove users.csv


def test_csv_unflattened_list_attributes():
    """Validates that list elements are flattened on save when readable flag is set to False on store()"""
    test_user = User("John Doe", 20, 75, ["Abby", "Mike", "Janice"])

    # Write test instance to spreadsheet
    with Spreadsheet('users.csv', User) as output_sheet:
        output_sheet.store(test_user)

    # Validate data saved properly
    with Spreadsheet('users.csv') as input_sheet:
        assert type(input_sheet) == _CSV_Spreadsheet
        _, instances = input_sheet.load("users")
        header = input_sheet.class_attributes

    # Check the list has been flattened
    assert instances[0].Family == ["Abby", "Mike", "Janice"]

    os.remove("users.csv")  # remove users.csv

def test_csv_unflattened_tuple_attributes():
    """Validates that tuple elements are flattened on save when readable flag is set to False on store()"""
    test_user = User("John Doe", 20, 75, ("Abby", "Mike", "Janice"))

    # Write test instance to spreadsheet
    with Spreadsheet('users.csv', User) as output_sheet:
        output_sheet.store(test_user)

    # Validate data saved properly
    with Spreadsheet('users.csv') as input_sheet:
        assert type(input_sheet) == _CSV_Spreadsheet
        _, instances = input_sheet.load("users")
        header = input_sheet.class_attributes

    assert instances[0].Family == ("Abby", "Mike", "Janice")

    os.remove("users.csv")  # remove users.csv


def test_csv_unflattened_dict_attributes():
    """Validates that dict elements are flattened on save when readable flag is set to False on store()"""
    test_user = User("John Doe", 20, 75, {"age": 21, "name": "Francis"})

    # Write test instance to spreadsheet
    with Spreadsheet('users.csv', User) as output_sheet:
        output_sheet.store(test_user)

    # Validate data saved properly
    with Spreadsheet('users.csv') as input_sheet:
        assert type(input_sheet) == _CSV_Spreadsheet
        _, instances = input_sheet.load("users")
        header = input_sheet.class_attributes

    assert instances[0].Family == {"age": "21", "name": "Francis"}

    os.remove("users.csv")  # remove users.csv


def test_csv_unmatched_instances():
    """Validates that an error is thrown when mismatched instance types are tried to be stored"""
    # Setup test instances
    leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

    philippine_eagle = Animal('Philippine Eagle', 'Threatened')

    # Try to use mismatching instances of class (User and Animal)
    try:
        with Spreadsheet('users.csv', User) as output_sheet:
            output_sheet.store(leopard_gecko, philippine_eagle)
        os.remove("users.csv")  # remove users.csv
        assert False # Test fails because it allows to store a different class than provided
    except ValueError:
        assert True


def test_csv_load_single_instance_no_identifier():
    """Validates that a single instance can be loaded from a sheet without a class identifier"""

    # Setup test instance
    leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

    # Write test instance to spreadsheet
    with Spreadsheet('animals.csv', Animal) as output_sheet:
        output_sheet.store(leopard_gecko)

    with Spreadsheet("animals.csv") as loaded_sheet:
        animals, instances = loaded_sheet.load("animals")

    assert animals.__name__ == "animals"
    assert type(animals) == type(namedtuple("animals", ["name", "conservation_status"]))
    assert animals._make
    assert instances[0].name == "Leopard Gecko"
    assert instances[0].conservation_status == "Least Concern"

    os.remove("animals.csv")  # remove animals.csv


def test_csv_load_single_instance_with_identifier():
    """Validates that a single instance can be loaded from a sheet with a class identifier"""

    # Setup test instance
    leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

    # Write test instance to spreadsheet
    with Spreadsheet('animals.csv', Animal) as output_sheet:
        output_sheet.store(leopard_gecko)

    with Spreadsheet("animals.csv", Animal) as loaded_sheet:
        animals, instances = loaded_sheet.load("animals")

    assert type(animals) == type(Animal)
    assert instances[0].name == "Leopard Gecko"
    assert instances[0].conservation_status == "Least Concern"

    os.remove("animals.csv")  # remove animals.csv


def test_csv_load_multi_instance_no_identifier():
    """Validates that multiple instance can be loaded from a sheet without a class identifier"""

    # Setup test instance
    leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

    philippine_eagle = Animal('Philippine Eagle', 'Threatened')

    # Write test instance to spreadsheet
    with Spreadsheet('animals.csv', Animal) as output_sheet:
        output_sheet.store(leopard_gecko, philippine_eagle)

    with Spreadsheet("animals.csv") as loaded_sheet:
        animals, instances = loaded_sheet.load("animals")

    assert animals.__name__ == "animals"
    assert type(animals) == type(namedtuple("animals", ["name", "conservation_status"]))
    assert animals._make
    assert len(instances) == 2
    assert instances[0].name == "Leopard Gecko"
    assert instances[0].conservation_status == "Least Concern"
    assert instances[1].name == "Philippine Eagle"
    assert instances[1].conservation_status == "Threatened"

    os.remove("animals.csv")  # remove animals.csv


def test_csv_load_multi_instance_with_identifier():
    """Validates that multiple instance can be loaded from a sheet with a class identifier"""

    # Setup test instance
    leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

    philippine_eagle = Animal('Philippine Eagle', 'Threatened')

    # Write test instance to spreadsheet
    with Spreadsheet('animals.csv', Animal) as output_sheet:
        output_sheet.store(leopard_gecko, philippine_eagle)

    with Spreadsheet("animals.csv") as loaded_sheet:
        animals, instances = loaded_sheet.load("animals")

    assert type(animals) == type(Animal)
    assert instances[0].name == "Leopard Gecko"
    assert instances[0].conservation_status == "Least Concern"
    assert instances[1].name == "Philippine Eagle"
    assert instances[1].conservation_status == "Threatened"

    os.remove("animals.csv")  # remove animals.csv


def test_csv_store_iterable_instances_no_identifier():
    """Validates that an iterable of instances can be loaded from a sheet with no identifier"""
    instances = []
    ranstring = lambda: ''.join(random.choices(string.ascii_uppercase, k=10)) # Generates a random 10 character string
    instances.append(User("John Doe", 20, 75, ["Abby", "Mike", "Janice"]))
    for i in range(1000):
        instances.append(User(ranstring(), random.randint(12,100), random.randint(75,400), [ranstring(), ranstring(), ranstring()]))

    with Spreadsheet('users.csv', User) as output_sheet:
        output_sheet.store(instances)

    with Spreadsheet("users.csv") as loaded_sheet:
        users, instances = loaded_sheet.load("users")

    assert users.__name__ == "users"
    assert type(users) == type(namedtuple("users", ["name", "age", "weight", "family"]))
    assert users._make
    assert instances[0].Name == "John Doe"
    assert instances[0].Age == 20
    assert instances[0].Weight == 75
    assert instances[0].Family == ["Abby", "Mike", "Janice"]
    assert len(instances) == 1001

    os.remove("users.csv")  # remove animals.csv

